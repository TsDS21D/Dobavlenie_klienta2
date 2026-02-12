"""
views.py
View-функции Django - обработчики HTTP запросов.
ДОБАВЛЕНЫ ФУНКЦИИ ДЛЯ РЕЗЕРВНОГО КОПИРОВАНИЯ И ЭКСПОРТА В EXCEL.

ИСПРАВЛЕНО:
- Добавлен импорт `timezone` из `django.utils`.
- Добавлен импорт `django` для получения версии.
- Все неопределённые имена устранены.
"""

from django.shortcuts import render, redirect
from django.views.decorators.cache import never_cache
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Order, Client
from .forms import LoginForm

# ===== НОВЫЕ ИМПОРТЫ (ДЛЯ ЭКСПОРТА И РЕЗЕРВИРОВАНИЯ) =====
import json
import django                                # для получения версии Django
from django.http import HttpResponse, JsonResponse
from django.core import serializers
from django.db import connection, transaction
from django.utils import timezone           # для работы с часовыми поясами
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


def login_view(request):
    """Представление для входа пользователя в систему."""
    
    if request.user.is_authenticated:
        return redirect('index')
    
    if request.method == 'POST':
        form = LoginForm(data=request.POST)
        
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            
            user = authenticate(request, username=username, password=password)
            
            if user is not None:
                login(request, user)
                messages.success(request, f'Добро пожаловать, {user.username}!')
                return redirect('index')
            else:
                messages.error(request, 'Неверное имя пользователя или пароль.')
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    
    else:
        form = LoginForm()
    
    return render(request, 'counter/login.html', {'form': form, 'title': 'Вход в систему'})


def logout_view(request):
    """Представление для выхода пользователя из системы."""
    
    logout(request)
    messages.info(request, 'Вы успешно вышли из системы.')
    
    return redirect('login')


@login_required(login_url='/login/')
@never_cache
def index(request):
    """Главная страница системы управления заказами."""
    
    orders = Order.objects.all()
    
    response = render(request, 'counter/index.html', {
        'orders': orders,
        'user': request.user,
    })
    
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response


# ========== НОВЫЕ ФУНКЦИИ ==========

@login_required
def export_orders_excel(request):
    """
    Экспорт всех заказов в файл Excel (.xlsx).
    Доступно только авторизованным пользователям.
    """
    # 1. Получаем все заказы из базы данных, сортируем по дате создания (новые сверху)
    orders = Order.objects.all().order_by('-created_at')
    
    # 2. Создаем новую книгу Excel
    wb = Workbook()
    
    # 3. Выбираем активный лист (по умолчанию создаётся один лист)
    ws = wb.active
    
    # 4. Устанавливаем название листа
    ws.title = "Заказы"
    
    # --- Определяем стили для оформления ---
    # Шрифт заголовков: жирный, размер 12, белый
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    
    # Заливка заголовков: фирменный зелёный цвет
    header_fill = PatternFill(start_color='0B8661', end_color='0B8661', fill_type='solid')
    
    # Выравнивание: по центру по горизонтали и вертикали, перенос текста
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Границы ячеек: тонкая серая линия
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # 5. Заголовки колонок (будет первая строка)
    headers = [
        '№ заказа',
        'Клиент',
        'Описание',
        'Дата готовности',
        'Статус',
        'Дата создания',
        'Телефон',
        'Email',
        'ЭДО'
    ]
    
    # 6. Записываем заголовки в первую строку (столбцы A, B, C, ...)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # 7. Заполняем данными (начиная со второй строки)
    for row_num, order in enumerate(orders, 2):
        # Получаем информацию о клиенте (если есть)
        client = order.client
        client_name = client.name if client else order.customer_name or "—"
        phone = client.phone if client else ""
        email = client.email if client else ""
        uses_edo = "Да" if (client and client.uses_edo) else "Нет"
        
        # Форматируем даты в московском времени (как в интерфейсе)
        # ready_datetime уже хранится в UTC, но timezone.localtime() конвертирует в текущий часовой пояс.
        ready_dt_local = timezone.localtime(order.ready_datetime)
        created_dt_local = timezone.localtime(order.created_at)
        
        ready_str = ready_dt_local.strftime('%d.%m.%Y %H:%M')
        created_str = created_dt_local.strftime('%d.%m.%Y %H:%M')
        
        # Значения ячеек в том же порядке, что и заголовки
        row_data = [
            f"{order.order_number:04d}",
            client_name,
            order.description,
            ready_str,
            order.get_status_display_name(),
            created_str,
            phone,
            email,
            uses_edo
        ]
        
        # Записываем данные в строку
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
    
    # 8. Автоматическая настройка ширины колонок
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # буква колонки (A, B, C...)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Устанавливаем ширину = максимальная длина + 2 (некоторый запас)
        adjusted_width = min(max_length + 2, 50)  # но не более 50 символов
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # 9. Формируем HTTP-ответ с Excel-файлом
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Заголовок Content-Disposition указывает браузеру, что это вложение и нужно скачать
    response['Content-Disposition'] = f'attachment; filename=orders_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # 10. Сохраняем книгу в response (wb.save принимает file-like объект)
    wb.save(response)
    
    return response


@login_required
def backup_download(request):
    """
    Скачать резервную копию базы данных (все клиенты и все заказы) в формате JSON.
    Используется встроенная сериализация Django.
    """
    # 1. Получаем все объекты моделей Client и Order
    clients = Client.objects.all()
    orders = Order.objects.all()
    
    # 2. Сериализуем их в JSON-строку
    #    serializers.serialize() возвращает строку в формате JSON,
    #    содержащую все поля и информацию о модели.
    data_clients = serializers.serialize('json', clients)
    data_orders = serializers.serialize('json', orders)
    
    # 3. Объединяем в один словарь, чтобы было удобно при восстановлении
    backup_data = {
        'clients': json.loads(data_clients),  # превращаем JSON-строку в объект Python
        'orders': json.loads(data_orders),
        'export_date': datetime.now().isoformat(),  # метка времени экспорта
        'django_version': django.get_version(),      # версия Django (для совместимости)
    }
    
    # 4. Превращаем обратно в JSON с отступами для читаемости
    backup_json = json.dumps(backup_data, ensure_ascii=False, indent=2)
    
    # 5. Создаём HTTP-ответ с файлом
    response = HttpResponse(backup_json, content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename=printshop_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
    
    return response


@login_required
@transaction.atomic  # Все операции внутри одной транзакции: либо всё выполнится, либо ничего
def backup_upload(request):
    """
    Восстановление базы данных из загруженного JSON-файла.
    Ожидается POST-запрос с полем 'backup_file', содержащим файл резервной копии.
    """
    if request.method != 'POST':
        # Если метод не POST, возвращаем ошибку 405 (Method Not Allowed)
        return JsonResponse({'success': False, 'error': 'Метод не разрешён'}, status=405)
    
    # 1. Проверяем, передан ли файл
    if 'backup_file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'Файл не загружен'}, status=400)
    
    backup_file = request.FILES['backup_file']
    
    # 2. Читаем содержимое файла и декодируем из JSON
    try:
        file_content = backup_file.read().decode('utf-8')
        backup_data = json.loads(file_content)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Ошибка чтения JSON: {str(e)}'}, status=400)
    
    # 3. Проверяем структуру файла: должны присутствовать ключи 'clients' и 'orders'
    if 'clients' not in backup_data or 'orders' not in backup_data:
        return JsonResponse({'success': False, 'error': 'Неверный формат резервной копии (отсутствуют clients или orders)'}, status=400)
    
    try:
        # 4. Очищаем существующие данные в правильном порядке:
        #    сначала удаляем заказы (т.к. они ссылаются на клиентов),
        #    потом удаляем клиентов.
        Order.objects.all().delete()
        Client.objects.all().delete()
        
        # 5. Сбрасываем последовательности автоинкремента для обеих таблиц,
        #    чтобы новые записи не конфликтовали со старыми ID, которые мы сейчас вставим.
        #    Для PostgreSQL используем raw SQL.
        with connection.cursor() as cursor:
            # Сброс последовательности для таблицы counter_client (поле id)
            cursor.execute("SELECT setval(pg_get_serial_sequence('counter_client','id'), COALESCE((SELECT MAX(id) FROM counter_client), 1), false);")
            # Сброс последовательности для таблицы counter_order (поле order_number)
            cursor.execute("SELECT setval(pg_get_serial_sequence('counter_order','order_number'), COALESCE((SELECT MAX(order_number) FROM counter_order), 1), false);")
        
        # 6. Десериализуем и сохраняем клиентов
        #    backup_data['clients'] - это список объектов в формате, который возвращает serializers.serialize
        #    deserialize() возвращает генератор объектов DeserializedObject,
        #    у каждого вызываем .save() для сохранения в БД.
        client_objects = []
        for client_obj in serializers.deserialize('json', json.dumps(backup_data['clients'])):
            client_objects.append(client_obj)
            client_obj.save()  # сохраняем с теми же primary key, что были в дампе
        
        # 7. Десериализуем и сохраняем заказы
        order_objects = []
        for order_obj in serializers.deserialize('json', json.dumps(backup_data['orders'])):
            order_objects.append(order_obj)
            order_obj.save()
        
        # 8. Ещё раз сбрасываем последовательности, чтобы они учитывали максимальные значения
        with connection.cursor() as cursor:
            cursor.execute("SELECT setval(pg_get_serial_sequence('counter_client','id'), COALESCE((SELECT MAX(id) FROM counter_client), 1), true);")
            cursor.execute("SELECT setval(pg_get_serial_sequence('counter_order','order_number'), COALESCE((SELECT MAX(order_number) FROM counter_order), 1), true);")
        
        # 9. Возвращаем успешный JSON-ответ
        #    Фронтенд может перезагрузить страницу после этого ответа.
        return JsonResponse({
            'success': True,
            'message': f'База данных успешно восстановлена. Загружено клиентов: {len(client_objects)}, заказов: {len(order_objects)}.'
        })
        
    except Exception as e:
        # В случае любой ошибки транзакция будет откачена автоматически (благодаря @transaction.atomic)
        return JsonResponse({'success': False, 'error': f'Ошибка при восстановлении: {str(e)}'}, status=500)